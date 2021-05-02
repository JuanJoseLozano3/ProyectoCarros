import tkinter as tk
from tkinter import *
from PIL import ImageTk, Image
import random
from tkinter import messagebox as MessageBox
from openpyxl import load_workbook
import pandas as pd
import numpy as np
from operator import itemgetter

#Creacion ventana inicial

root = tk.Tk()
root.title("Juego")
nombresjugadores=[]


#Jugadores

def asignarnombres():
    global k
    global nombresjugadores
    global root
    k=tk.StringVar()
    nombres = tk.Label(root, text=str("Nombre jugador " +str(len(nombresjugadores)+1)), font=('courier', 10, 'bold'), justify=tk.LEFT)
    nombres.grid(row=1,column=1)
    varf = tk.Entry(root, textvariable=k)
    varf.grid(row=2,column=1)
    v = tk.Button(root,text="Listo", command=asignom)
    v.grid(row=3,column=1)

def asignom():
    global k
    nombresjugadores.append(k.get())
    Limpiar()

def inijugadores():
    global njugadores
    global podio
    global podios
    global nombresjugadores
    global root
    nombresjugadores=[]
    podio=[]
    podios=[]
    njug = tk.Label(root, text="Numero de Jugadores", font=('courier', 20, 'bold'), justify=tk.LEFT)
    njugas = tk.Label(root, text="Maximo 20", font=('courier', 10, 'bold'), justify=tk.LEFT)
    njug.grid(row=0,column=1)
    njugas.grid(row=1,column=1)

    njugadores=tk.StringVar()
    njugador = tk.Entry(root, textvariable=njugadores, font=('courier', 10, 'bold'), bg='white')
    njugador.grid(row=2,column=1)

    Listojug = tk.Button(root,text="Aceptar", command=Limpiar)
    Listojug.grid(row=3, column=1)

    Verpodio = tk.Button(root,text="Ver podio", command=verpodio)
    Verpodio.grid(row=4, column=1)


def Limpiar():
    global nombrejugadores
    global root
    root.destroy()
    root = tk.Tk()
    if(len(nombresjugadores)<int(njugadores.get())):
        asignarnombres()
    else:
        carros()

def volv():
    global root
    root.destroy()
    root = tk.Tk()
    inijugadores()




#Carros

def carros():
    global jugad
    global ganador
    global njugadores
    global nombresjugadores
    global img
    global flag2
    global puestos
    global ganado
    global puntos
    global podio
    global podios
    global i
    global root
    puestos=0
    img=[]
    imgmeta=[]
    pos=[]
    posmeta=[]
    puntos=[]
    ganador=[]
    ganado=[]
    i=0
    flag2=0
    while(i<(int(njugadores.get()))):
        img.append(PhotoImage(file='carro'+str(i+1)+'.png'))
        imgmeta.append(PhotoImage(file='meta.png'))
        img[i] = img[i].zoom(4)
        img[i] = img[i].subsample(32)
        imgmeta[i] = imgmeta[i].zoom(4)
        imgmeta[i] = imgmeta[i].subsample(32) 
        pos.append(tk.Label(root,image=img[i]))
        pos[i].grid(row=(i+1),column=0)
        posmeta.append(tk.Label(root,image=imgmeta[i]))
        posmeta[i].grid(row=(i+1),column=25)
        puntos.append(0)
        i+=1
    jugad=1
    tk.Label(root, text="Turno", font=('courier', 10, 'bold'), justify=tk.LEFT).grid(row=i+1,column=11)
    tk.Label(root, text="del", font=('courier', 10, 'bold'), justify=tk.LEFT).grid(row=i+1,column=12)
    tk.Label(root, text="jugador", font=('courier', 10, 'bold'), justify=tk.LEFT).grid(row=i+1,column=13)
    tk.Label(root, text='       ', font=('courier', 10, 'bold'), justify=tk.LEFT).grid(row=i+1,column=14)
    tk.Label(root, text=str(nombresjugadores[0]), font=('courier', 10, 'bold'), justify=tk.LEFT).grid(row=i+1,column=14)
    tk.Button(root, text="Tirar", command=dados).grid(row=i+1,column=15)
    tk.Button(root, text="Fin", command=fin).grid(row=i+1,column=16)
    pista()
    root.mainloop()



#Dados

def dados():
    global jugad
    global noimbresjugadores
    global ganador
    global i
    global root
    tirardados()
    jugad+=1
    if(jugad>int(njugadores.get())):
        jugad=1
    while(jugad<=int(njugadores.get())):
        if(jugad in ganador):
            jugad+=1
            if(jugad>int(njugadores.get())):
                jugad=1
        else:
            break
    tk.Label(root, text=str('       '), font=('courier', 10, 'bold'), justify=tk.LEFT).grid(row=i+1,column=14)
    tk.Label(root, text=str(nombresjugadores[jugad-1]), font=('courier', 10, 'bold'), justify=tk.LEFT).grid(row=i+1,column=14)


def tirardados():
    global dado
    global nombresjugadores
    global jugad
    global root
    dado=random.randint(1,6)
    mensaje=str("Jugador "+str(nombresjugadores[jugad-1])+" obtuvo "+str(dado)+" puntos ")
    MessageBox.showinfo("Tiro dados", mensaje) 
    carril()




#Pista
def pista():
    global root
    i=0
    while(i<=25):
        tk.Label(root, text=str(str(i*100)+'M'), font=('courier', 10, 'bold'), justify=tk.LEFT).grid(row=0,column=i)
        i+=1
    



#Carril

def carril():
    global img
    global poscarro
    global puntos
    global jugad
    global poscarro
    global root
    im=(PhotoImage(file='limpiar.png'))
    imagencarro=img[int(jugad)-1]
    poscarroant=puntos[jugad-1]
    carrito=tk.Label(root,image=im).grid(row=(jugad),column=poscarroant)
    poscarro=puntos[jugad-1]+dado
    puntos[jugad-1]=poscarro
    poscarrito=tk.Label(root,image=imagencarro).grid(row=(jugad),column=poscarro)
    final()

def final():
    global jugad
    global flag2
    global puntos
    global ganador
    global ganado
    if(puntos[jugad-1]>24 and flag2==0 and (len(ganador))<3):
        if(int(njugadores.get())==1):
            ganador.append(jugad)
            ganado.append(str(nombresjugadores[jugad-1]))
            ganadores()
        elif(int(njugadores.get())==2):
            ganador.append(jugad)
            ganado.append(str(nombresjugadores[jugad-1]))
            flag2=1
        else:
            ganador.append(jugad)
            ganado.append(str(nombresjugadores[jugad-1]))
    else:
        ganadores()
    if(len(ganador)==3):
        ganadores()


#Ganadores
def ganadores():
    global ganador
    global ganado
    global podio
    global podios
    global flag2
    global root
    if(int(njugadores.get())==1):
        MessageBox.showinfo("Galardones", "Primer puesto para el jugador "+str(ganado[0]))
        podio.append(ganador[0])
        podios.append(ganado[0])
        Limpiar()
    elif(int(njugadores.get())==2 and flag2==1):
        MessageBox.showinfo("Galardones", "Primer puesto para el jugador "+str(ganado[0]))
        MessageBox.showinfo("Galardones", "Segundo puesto para el jugador "+str(ganado[1]))
        podio.append(ganador[0])
        podios.append(ganado[0])
        podio.append(ganador[1])
        podios.append(ganado[1])
        Limpiar()
    elif(len(ganador)==3):
        MessageBox.showinfo("Galardones", "Primer puesto para el jugador "+str(ganado[0]))
        MessageBox.showinfo("Galardones", "Segundo puesto para el jugador "+str(ganado[1]))
        MessageBox.showinfo("Galardones", "Tercer puesto para el jugador "+str(ganado[2]))
        podio.append(ganador[0])
        podios.append(ganado[0])
        podio.append(ganador[1])
        podios.append(ganado[1])
        podio.append(ganador[2])
        podios.append(ganado[2])
        Limpiar()

#Podio

def finals():
    global primeron
    global segundon
    global terceron
    global podio
    wb = load_workbook('puestos.xlsx')
    hoja = wb.active
    ws1=wb['Hoja1']
    max=ws1.max_row
    maxs=ws1.max_row
    i=0
    while (i<len(primero)):
        hoja['A%d' % ((i+1+maxs))] = max
        hoja['B%d' % ((i+1+maxs))] = str(str(primeron[i]))
        hoja['C%d' % ((i+1+maxs))] = str(str(segundon[i]))
        hoja['D%d' % ((i+1+maxs))] = str(str(terceron[i]))
        max+=1
        i+=1
    # Guardo la hoja
    wb.save('puestos.xlsx')
    exit()
 

def verpodio():
    LimpiarT()
    wb = load_workbook('puestos.xlsx')
    hoja = wb.active
    ws1=wb['Hoja1']
    maxs=ws1.max_row
    puesti = tk.Label(root, text="Primer puesto", font=('courier', 10, 'bold'), justify=tk.LEFT)
    puesti.grid(row=0,column=1)
    puesti = tk.Label(root, text="Partidas ganadas", font=('courier', 10, 'bold'), justify=tk.LEFT)
    puesti.grid(row=0,column=2)
    pus=[]
    i=2
    while(i<=maxs):
        pus.append(hoja[str('B'+str(i))].value)
        i+=1
    pas=pd.unique(pus)
    pod=np.array(pus)
    vals=[]
    i=0
    while(i<len(pas)):
        vals.append(np.count_nonzero(pod == pas[i]))
        i+=1
    numbers_sort = sorted(enumerate(vals), key=itemgetter(1),  reverse=True)
    index, value = numbers_sort[0]
    puesti1 = tk.Label(root, text=str(pas[index]), font=('courier', 10, 'bold'), justify=tk.LEFT)
    puesti1.grid(row=2,column=1)
    puesti1v = tk.Label(root, text=str(value), font=('courier', 10, 'bold'), justify=tk.LEFT)
    puesti1v.grid(row=2,column=2)
    numbers_sort = sorted(enumerate(vals), key=itemgetter(1),  reverse=True)
    index, value = numbers_sort[1]
    puesti2 = tk.Label(root, text=str(pas[index]), font=('courier', 10, 'bold'), justify=tk.LEFT)
    puesti2.grid(row=3,column=1)
    puesti2v = tk.Label(root, text=str(value), font=('courier', 10, 'bold'), justify=tk.LEFT)
    puesti2v.grid(row=3,column=2)
    numbers_sort = sorted(enumerate(vals), key=itemgetter(1),  reverse=True)
    index, value = numbers_sort[2]
    puesti3 = tk.Label(root, text=str(pas[index]), font=('courier', 10, 'bold'), justify=tk.LEFT)
    puesti3.grid(row=4,column=1)
    puesti3v = tk.Label(root, text=str(value), font=('courier', 10, 'bold'), justify=tk.LEFT)
    puesti3v.grid(row=4,column=2)

    Verpodio = tk.Button(root,text="Volver", command=volv)
    Verpodio.grid(row=5, column=1)


#Limpiar
def LimpiarT():
    global root
    root.destroy()
    root = tk.Tk()


#Fin
def fin():
    global podio
    global podios
    global primero
    global segundo
    global tercero
    global primeron
    global segundon
    global terceron
    global root
    LimpiarT()
    primero=[]
    segundo=[]
    tercero=[]
    primeron=[]
    segundon=[]
    terceron=[]
    tk.Label(root, text='Partida', font=('courier', 10, 'bold'), justify=tk.LEFT).grid(row=0,column=0)
    tk.Label(root, text='Primer puesto', font=('courier', 10, 'bold'), justify=tk.LEFT).grid(row=0,column=1)
    tk.Label(root, text='Segundo puesto', font=('courier', 10, 'bold'), justify=tk.LEFT).grid(row=0,column=2)
    tk.Label(root, text='Tercer puesto', font=('courier', 10, 'bold'), justify=tk.LEFT).grid(row=0,column=3)
    i=0
    j=0
    while(i<len(podio)):
        j+=1
        tk.Label(root, text=str(j), font=('courier', 10, 'bold'), justify=tk.LEFT).grid(row=j,column=0)
        tk.Label(root, text=str('jugador '+str(podios[i])), font=('courier', 10, 'bold'), justify=tk.LEFT).grid(row=j,column=1)
        primero.append((podio[i]))
        primeron.append((podios[i]))
        tk.Label(root, text=str('jugador '+str(podios[i+1])), font=('courier', 10, 'bold'), justify=tk.LEFT).grid(row=j,column=2)
        segundo.append((podio[i+1]))
        segundon.append((podios[i+1]))
        tk.Label(root, text=str('jugador '+str(podios[i+2])), font=('courier', 10, 'bold'), justify=tk.LEFT).grid(row=j,column=3)
        tercero.append((podio[i+2]))
        terceron.append((podios[i+2]))
        i+=3
    tk.Button(root, text='Contador', command=contador).grid(row=j+1,column=1)
    tk.Button(root, text='Final', command=finals).grid(row=j+1,column=2)

def contador():
    global primero
    global segundo
    global tercero
    global root
    LimpiarT()
    i=1
    c=0
    co=[]
    tk.Label(root, text='Primer puesto', font=('courier', 10, 'bold'), justify=tk.LEFT).grid(row=0,column=1)
    tk.Label(root, text='Segundo puesto', font=('courier', 10, 'bold'), justify=tk.LEFT).grid(row=0,column=2)
    tk.Label(root, text='Tercer puesto', font=('courier', 10, 'bold'), justify=tk.LEFT).grid(row=0,column=3)
    while(i<=int(njugadores.get())):
        co.append([])
        tk.Label(root, text=str(nombresjugadores[i-1]), font=('courier', 10, 'bold'), justify=tk.LEFT).grid(row=i,column=0)
        j=0
        while(j<len(primero)):
            if(i==primero[j]):
                c+=1     
            j+=1
        co[i-1].append(c)
        tk.Label(root, text=co[i-1][0], font=('courier', 10, 'bold'), justify=tk.LEFT).grid(row=i,column=1)
        c=0
        j=0
        while(j<len(segundo)):
            
            if(i==segundo[j]):
                c+=1     
            j+=1
        co[i-1].append(c)
        tk.Label(root, text=co[i-1][1], font=('courier', 10, 'bold'), justify=tk.LEFT).grid(row=i,column=2)
        c=0
        j=0
        while(j<len(tercero)):
            if(i==tercero[j]):
                c+=1     
            j+=1
        co[i-1].append(c)
        tk.Label(root, text=co[i-1][2], font=('courier', 10, 'bold'), justify=tk.LEFT).grid(row=i,column=3)
        c=0
        i+=1
    tk.Button(root, text='Final', command=finals).grid(row=21,column=j+1)

 


inijugadores()
root.mainloop()
